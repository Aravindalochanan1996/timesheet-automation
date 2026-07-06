from openpyxl import Workbook
from datetime import datetime
import calendar

from app.report.formatter import format_sheet, highlight_remarks
from app.report.summary_generator import (
    generate_summary_sheet,
    calculate_total_hours
)

from app.report.working_days_calculator import calculate_metrics
from app.report.leave_metrics import extract_leave_metrics

# ✅ Correct Import
from app.report.timesheet_enrolled import generate_timesheet_enrolled_sheet

from app.config.settings import EMPLOYEES_FILE


def generate_excel(records, output_file):

    wb = Workbook()

    grouped = {}

    # =========================
    # Group by month
    # =========================
    for r in records:
        grouped.setdefault(r["month"], []).append(r)

    # =========================
    # Sort months latest first
    # =========================
    sorted_months = sorted(
        grouped.keys(),
        key=lambda m: datetime.strptime(m, "%b-%Y"),
        reverse=True
    )

    # =========================
    # Generate Month Sheets
    # =========================
    for month in sorted_months:

        recs = sorted(
            grouped[month],
            key=lambda r: str(r["name"]).strip().lower()
        )

        month_name, year = month.split("-")

        ws = wb.create_sheet(title=f"{month_name} {year}")

        month_index = datetime.strptime(month_name, "%b").month

        days = calendar.monthrange(
            int(year),
            month_index
        )[1]

        # =========================
        # Header
        # =========================
        employee_count = len(recs)

        header = [
            f"Employee Name ({employee_count})",
            "Employee ID",
            "Month"
        ]

        for d in range(1, days + 1):
            header += [
                f"Day {d} Hours",
                f"Day {d} Remarks"
            ]

        header.extend([
            "Total Hours",
            "Total Working Days",
            "Unpaid Leave Count",
            "Public Holiday Dates",
            "Unpaid Leave Dates",
            "Final Work Hours",
            "Report Generated Date"
        ])

        ws.append(header)

        # =========================
        # Populate rows
        # =========================
        for r in recs:

            row = [
                r["name"],
                r["emp_id"],
                r["month"]
            ]

            # Daily entries
            for d in range(1, days + 1):

                key = str(d).zfill(2)

                entry = r["entries"].get(key, {})

                row.append(entry.get("hours", ""))
                row.append(entry.get("remarks", ""))

            # Metrics
            working_days, unpaid_days = calculate_metrics(
                r["entries"],
                r["month"]
            )

            total_hours = calculate_total_hours(r["entries"])

            public_holidays, unpaid_count, unpaid_dates = (
                extract_leave_metrics(
                    r["entries"],
                    r["month"]
                )
            )

            final_hours = total_hours - (unpaid_count * 8)

            # Append metrics
            row.extend([
                total_hours,
                working_days,
                unpaid_count,
                public_holidays,
                unpaid_dates,
                final_hours,
                r.get("report_generated", "")
            ])

            ws.append(row)

    # =========================
    # Remove default sheet
    # =========================
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # =========================
    # Summary Sheet
    # =========================
    generate_summary_sheet(wb, records)

    # =========================
    # Timesheet Enrolled Sheet
    # =========================
    generate_timesheet_enrolled_sheet(
        wb,
        records,
        EMPLOYEES_FILE
    )

    # =========================
    # Sort sheets
    # =========================
    sheet_priority = {
        "Summary": 0,
        "Timesheet Enrolled": 1
    }

    wb._sheets.sort(
        key=lambda ws: (
            sheet_priority.get(ws.title, 2),
            -datetime.strptime(ws.title, "%b %Y").timestamp()
            if ws.title not in sheet_priority
            else 0
        )
    )

    # =========================
    # Apply formatting
    # =========================
    for sheet in wb.worksheets:

        if sheet.title != "Summary":
            highlight_remarks(sheet)

        format_sheet(sheet)

    # =========================
    # Save workbook
    # =========================
    wb.save(output_file)