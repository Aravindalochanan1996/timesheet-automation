from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==============================
# 🎨 Styles
# ==============================

# Highlight for remarks
REMARK_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

HEADER_FONT = Font(bold=True)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADER_FILL = PatternFill(start_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="C6E0B4", fill_type="solid")


# ==============================
# 🎯 Highlight Remarks
# ==============================

def highlight_remarks(ws):
    header = [cell.value for cell in ws[1]]

    remark_cols = [i + 1 for i, col in enumerate(header) if "Remarks" in str(col)]

    for row in ws.iter_rows(min_row=2):
        for col_idx in remark_cols:
            cell = row[col_idx - 1]

            if cell.value and str(cell.value).strip():
                cell.fill = REMARK_FILL


# ==============================
# 📊 Format Sheet
# ==============================

def format_sheet(ws):
    # ==============================
    # Freeze + View Fix
    # ==============================
    ws.freeze_panes = "A2"
    ws.sheet_view.topLeftCell = "A1"   # ✅ Fix header visibility
    # ws.auto_filter.ref = ws.dimensions  # ✅ Enable filter

    # ==============================
    # Identify columns
    # ==============================
    header = [cell.value for cell in ws[1]]
    total_col_index = len(header)
    
    # Fix Excel rendering issue
    ws.sheet_view.zoomScale = 100

    # ✅ Auto Adjust header height
    max_lines = 1

    for cell in ws[1]:
        if cell.value:
            text = str(cell.value)
            lines = text.count("\n") + 1
            max_lines = max(max_lines, lines)

    # Approx height calculation
    ws.row_dimensions[1].height = max(20, max_lines * 15)

    # ==============================
    # Column Width (FIXED LOGIC)
    # ==============================
    for col_idx, col_name in enumerate(header, start=1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        col_name = str(col_name)

        if "Remarks" in col_name:
            ws.column_dimensions[column_letter].width = 28

        elif "Final Work Hours" in col_name:
            ws.column_dimensions[column_letter].width = 22   

        elif "Report Generated Date" in col_name:
            ws.column_dimensions[column_letter].width = 26   

        elif "Unpaid Leave Dates" in col_name:
            ws.column_dimensions[column_letter].width = 25

        elif "Public Holidays" in col_name:
            ws.column_dimensions[column_letter].width = 22

        elif "Hours" in col_name:
            ws.column_dimensions[column_letter].width = 14

        elif "Date" in col_name:
            ws.column_dimensions[column_letter].width = 20

        elif col_name in ["Name", "Emp ID", "Month"]:
            ws.column_dimensions[column_letter].width = 20

        else:
            ws.column_dimensions[column_letter].width = 18

    # ==============================
    # Apply styles
    # ==============================
    for row in ws.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER_ALIGN
            else:
                header_value = str(ws.cell(row=1, column=cell.column).value)

                if "Remarks" in header_value:
                    cell.alignment = LEFT_ALIGN
                else:
                    cell.alignment = CENTER_ALIGN

            cell.border = BORDER

    # ==============================
    # Highlight Total Column
    # ==============================
    for row in ws.iter_rows(min_row=2):
        total_cell = row[total_col_index - 1]
        total_cell.fill = TOTAL_FILL