from openpyxl.chart import BarChart, Reference


def convert_to_hours(time_str):
    try:
        if not time_str or time_str == "--:--":
            return 0
        h, m = time_str.split(":")
        return int(h) + int(m) / 60
    except:
        return 0


def add_unpaid_leave_chart(ws):
    data = {}

    header = [cell.value for cell in ws[1]]

    hour_cols = []
    remark_cols = []

    for i, col in enumerate(header):
        if "Hours" in str(col) and "Total" not in str(col):
            hour_cols.append(i + 1)
        elif "Remarks" in str(col):
            remark_cols.append(i + 1)

    for row in ws.iter_rows(min_row=2):
        name = str(row[0].value).strip()

        if name not in data:
            data[name] = 0

        for h_col, r_col in zip(hour_cols, remark_cols):
            hours = convert_to_hours(row[h_col - 1].value)
            remarks = str(row[r_col - 1].value).lower()

            if "unpaid" in remarks:
                data[name] += hours

    # -------- Create summary table --------
    start_row = ws.max_row + 3

    ws.cell(row=start_row, column=1, value="Employee")
    ws.cell(row=start_row, column=2, value="Unpaid Leave Hours")

    for i, (name, hours) in enumerate(data.items(), start=1):
        ws.cell(row=start_row + i, column=1, value=name)
        ws.cell(row=start_row + i, column=2, value=hours)

    # -------- ✅ CREATE NEW CHART HERE --------
    chart = BarChart()

    chart.title = "Unpaid Leave per Employee"
    chart.y_axis.title = "Hours"
    chart.x_axis.title = "Employee"

    data_ref = Reference(
        ws,
        min_col=2,
        min_row=start_row,
        max_row=start_row + len(data)
    )

    cats_ref = Reference(
        ws,
        min_col=1,
        min_row=start_row + 1,
        max_row=start_row + len(data)
    )

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, f"E{start_row}")
    ws.add_chart(chart, "Z5")