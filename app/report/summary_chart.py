from openpyxl.chart import BarChart, Reference
from datetime import datetime


def is_weekend(year, month, day):
    try:
        dt = datetime(year, month, day)
        return dt.weekday() >= 5  # 5=Sat, 6=Sun
    except:
        return False


def add_summary_pivot_chart(ws, records):
    data = {}

    for r in records:
        name = r["name"].strip()
        month_str = r["month"]  # Mar-2026

        month_name, year = month_str.split("-")
        year = int(year)

        month_index = datetime.strptime(month_name, "%b").month

        if name not in data:
            data[name] = {"working": 0, "unpaid": 0}

        for day, entry in r["entries"].items():
            day_int = int(day)

            if is_weekend(year, month_index, day_int):
                continue

            hours = entry.get("hours", "")
            remarks = str(entry.get("remarks", "")).lower()

            if "unpaid" in remarks:
                data[name]["unpaid"] += 1

            elif hours not in ["", "--:--"]:
                data[name]["working"] += 1

    # -------- Create table --------
    start_row = ws.max_row + 5

    ws.cell(row=start_row, column=1, value="Employee")
    ws.cell(row=start_row, column=2, value="Working Days")
    ws.cell(row=start_row, column=3, value="Unpaid Leaves")

    for i, (name, val) in enumerate(data.items(), start=1):
        ws.cell(row=start_row + i, column=1, value=name)
        ws.cell(row=start_row + i, column=2, value=val["working"])
        ws.cell(row=start_row + i, column=3, value=val["unpaid"])

    # -------- Create chart --------
    chart = BarChart()
    chart.title = "Working Days vs Unpaid Leaves"
    chart.y_axis.title = "Days"
    chart.x_axis.title = "Employee"

    data_ref = Reference(
        ws,
        min_col=2,
        min_row=start_row,
        max_col=3,
        max_row=start_row + len(data)
    )

    cats_ref = Reference(
        ws,
        min_col=1,
        min_row=start_row + 1,
        max_row=start_row + len(data)
    )

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, f"E{start_row}")