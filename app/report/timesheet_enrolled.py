import pandas as pd

from datetime import datetime
from openpyxl.styles import PatternFill, Alignment
from app.report.working_days_calculator import calculate_metrics

def generate_timesheet_enrolled_sheet(wb, records, EMPLOYEES_FILE):

    # =========================
    # Green Fill for YES
    # =========================
    YES_FILL = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    NO_FILL = PatternFill(
        start_color="FFFFFF",
        end_color="FFFFFF",
        fill_type="solid"
    )

    # =========================
    # Read Employee Master
    # =========================
    df = pd.read_excel(EMPLOYEES_FILE)

    # Normalize headers
    df.columns = df.columns.str.strip().str.lower()

    employees = []

    for _, row in df.iterrows():

        employees.append({
            "name": (
                "-"
                if pd.isna(row.get("employee name"))
                else str(row.get("employee name")).strip().lower()
            ),

            "old_id": (
                "-"
                if pd.isna(row.get("old employee id"))
                else str(row.get("old employee id")).replace(".0", "").strip()
            ),

            "new_id": (
                "-"
                if pd.isna(row.get("new employee id"))
                else str(row.get("new employee id")).replace(".0", "").strip()
            ),

            "location": (
                "-"
                if pd.isna(row.get("location"))
                else str(row.get("location")).strip()
            )
        })

    # =========================
    # Sort Months Latest First
    # =========================
    months = sorted(
        list(set(r["month"] for r in records)),
        key=lambda m: datetime.strptime(m, "%b-%Y"),
        reverse=True
    )

    # =========================
    # Create Sheet
    # =========================
    ws = wb.create_sheet(title="Timesheet Enrolled")

    month_yes_count = {}

    for month in months:

        count = 0

        for emp in employees:

            matched = False

            for r in records:

                record_name = str(r["name"]).strip().lower()

                record_emp_id = (
                    str(r["emp_id"])
                    .replace(".0", "")
                    .strip()
                )

                if r["month"] != month:
                    continue

                old_match = (
                    emp.get("name", "") == record_name
                    and emp.get("old_id", "") == record_emp_id
                )

                new_match = (
                    emp.get("name", "") == record_name
                    and emp.get("new_id", "") == record_emp_id
                )

                if old_match or new_match:
                    matched = True
                    break

            if matched:
                count += 1

        month_yes_count[month] = count

    # Header
    header = [
        "Employee Name",
        "Old Employee ID",
        "New Employee ID",
        "Location Details"
    ]

    for month in months:

        count = month_yes_count.get(month, 0)

        header.append(
            f"{month}\n({count})"
        )

    ws.append(header)

    # =========================
    # Build lookup set
    # =========================
    lookup = {}

    for r in records:

        key = (
            str(r["name"]).strip().lower(),
            str(r["emp_id"]).replace(".0", "").strip(),
            r["month"]
        )

        working_days, unpaid_days = calculate_metrics(
            r["entries"],
            r["month"]
        )

        total_days = working_days + unpaid_days

        lookup[key] = total_days

    # =========================
    # Populate rows
    # =========================
    for emp in employees:

        row = [
            emp.get("name", "").title(),
            emp.get("old_id", ""),
            emp.get("new_id", ""),
            emp.get("location", "")
        ]

        for month in months:

            status = "NO"

            # Match with OLD ID
            key_old = (
                emp.get("name", ""),
                emp.get("old_id", ""),
                month
            )

            # Match with NEW ID
            key_new = (
                emp.get("name", ""),
                emp.get("new_id", ""),
                month
            )

            total_days = None

            if key_old in lookup:

                total_days = lookup[key_old]

            elif key_new in lookup:

                total_days = lookup[key_new]

            # Status Logic
            if total_days is not None:

                if total_days <= 10:
                    status = "PARTIAL"
                else:
                    status = "YES"

            row.append(status)

        # Add row
        ws.append(row)

        # Left align Employee Name column
        ws.cell(
            row=ws.max_row,
            column=1
        ).alignment = Alignment(horizontal="left")

        # =========================
        # Highlight YES / PARTIAL
        # =========================
        current_row = ws.max_row

        for col in range(5, len(header) + 1):

            cell = ws.cell(row=current_row, column=col)

            value = str(cell.value).strip().upper()

            if value == "YES":

                cell.fill = YES_FILL

            elif value == "PARTIAL":

                cell.fill = PatternFill(
                    start_color="FFF2CC",
                    end_color="FFF2CC",
                    fill_type="solid"
                )

            else:

                cell.fill = NO_FILL